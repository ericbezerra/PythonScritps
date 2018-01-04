## Author: Eric Bezerra
## Program: report.py
## Date: 03/01/2018
## Action: Create reports by wms system and
## send from email to clients
import requests as req
import openpyxl as xl
import re, bs4

class ReportBot():
	directory = r'C:\Users\eric.bezerra\Documents\Scripts\resource\\'
	tasks = r'C:\Users\eric.bezerra\Documents\Scripts\tasks\hosts.xlsx'
	keys = ['report', 'report123']
	hosts = []
	inds_graph = []
	memory_graph = []
	processor_graph = []
	tickets = []

	def __init__(self):
		print('Bot Started...')

## Read name of hosts
	def readHosts(self):
		

		wb = xl.load_workbook(self.tasks)
		sheet = wb.get_sheet_by_name('hosts')
		for i in range(1, sheet.max_row+1):
			self.hosts.append(sheet.cell(row=i, column=1).value)
				
## TODO: Colect Data from WMS and Files
## client = name of host
## start (list) = day, month, year, hour, minutes
## end (list) = day, month, year, hour, minutes
	def takeData(self, client, start, end):
		## Configuration of PNG
		link = 'http://10.130.254.10/wms/pnp4nagios/index.php/image?host='
		date = '&start='+str(start[1])+'%2F'+str(start[0])+'%2F'+str(start[2])+'+'+str(start[3])+'%3A'+str(start[4])+'+&end='+str(end[1])+'%2F'+str(end[0])+'%2F'+str(end[2])+'+'+str(end[3])+'%3A'+str(end[4])
		
		## Configuration of Graphs
		timeStart = '&smon='+str(start[1])+'&sday='+str(start[0])+'&syear='+str(start[2])+'&shour='+str(start[3])+'&smin='+str(start[4])+'&ssec=0'
		timeEnd = '&emon='+str(end[1])+'&eday='+str(end[0])+'&eyear='+str(end[2])+'&ehour='+str(end[3])+'&emin='+str(end[4])+'&esec=0'
		urlEnd = '&rpttimeperiod=&assumeinitialstates=yes&assumestateretention=yes&assumestatesduringnotrunning=yes&includesoftstates=yes&initialassumedservicestate=6&backtrack=4'
		
		## INDISPONIBILIDADE GRAFICO
		self.inds_graph = self.getTable('http://10.130.254.10/wms/nagios/cgi-bin/avail.cgi?show_log_entries=&host='+client+'&timeperiod=custom',timeStart,timeEnd,urlEnd)
		
		## TODO: INTERFACES GRAFICO

		## MEMORY GRAFICO
		self.memory_graph = self.getTable('http://10.130.254.10/wms/nagios/cgi-bin/avail.cgi?show_log_entries=&host='+client+'&service=Mem+used+Processor&timeperiod=custom',timeStart,timeEnd,urlEnd)

		## MEMORY PNG
		image = req.get(link+client+date+'&srv=Mem_used_Processor&theme=multisite&baseurl=..%2Fcheck_mk%2F', auth=(self.keys[0], self.keys[1]))
		img = open(self.directory+client+'_memory.png', 'wb')
		for chunk in image.iter_content(1000000):
			img.write(chunk)
		img.close()

		## PROCESSOR GRAFICO
		self.processor_graph = self.getTable('http://10.130.254.10/wms/nagios/cgi-bin/avail.cgi?show_log_entries=&host='+client+'&service=CPU+utilization&timeperiod=custom',timeStart,timeEnd,urlEnd)

		## PROCESSOR PNG
		image = req.get(link+client+date+'&srv=CPU_utilization&theme=multisite&baseurl=..%2Fcheck_mk%2F', auth=(self.keys[0], self.keys[1]))
		img = open(self.directory+client+'_processor.png', 'wb')
		for chunk in image.iter_content(1000000):
			img.write(chunk)
		img.close()

		## ICMP FULL PNG
		image = req.get(link+client+date+'+&srv=_HOST_&view=1&source=0', auth=(self.keys[0], self.keys[1]))
		img = open(self.directory+client+'_icmpfull.png', 'wb')
		for chunk in image.iter_content(1000000):
			img.write(chunk)
		img.close()

		## ICMP FIRST QUARTER PNG
		image = req.get(link+client+'&start='+str(start[1])+'%2F'+str(start[0])+'%2F'+str(start[2])+'+'+str(start[3])+'%3A'+str(start[4])+'+&end='+str(end[1])+'%2F'+str(end[0]-15)+'%2F'+str(end[2])+'+'+str(end[3])+'%3A'+str(end[4])+'+&srv=_HOST_&view=1&source=0', auth=(self.keys[0], self.keys[1]))
		img = open(self.directory+client+'_icmpstart.png', 'wb')
		for chunk in image.iter_content(1000000):
			img.write(chunk)
		img.close()

		## ICMP LAST QUARTER PNG
		image = req.get(link+client+'&start='+str(start[1])+'%2F'+str(start[0]+15)+'%2F'+str(start[2])+'+'+str(start[3])+'%3A'+str(start[4])+'+&end='+str(end[1])+'%2F'+str(end[0])+'%2F'+str(end[2])+'+'+str(end[3])+'%3A'+str(end[4])+'+&srv=_HOST_&view=1&source=0', auth=(self.keys[0], self.keys[1]))
		img = open(self.directory+client+'_icmpend.png', 'wb')
		for chunk in image.iter_content(1000000):
			img.write(chunk)
		img.close()

		## TODO: SERVICES CSV
		csv = req.get('http://10.130.254.10/wms/check_mk/view.py?selection=e1cd7915-e519-42ba-950e-7171589e2d9e&host='+client+'&view_name=host&st0=on&st1=on&st2=on&st3=on&stp=on&output_format=csv_export', auth=(self.keys[0], self.keys[1]))
		csv_file = open(self.directory+client+'_services.csv', 'wb')
		for chunk in csv.iter_content(1000000):
			csv_file.write(chunk)
		csv_file.close()

		## TODO: TICKETS OPEN GRAFICO
		wb = xl.load_workbook(r'C:\Users\eric.bezerra\Documents\Scripts\tasks\tickets.xlsx')
		sheet = wb.get_sheet_by_name('Tickets')
		for i in range(1, sheet.max_row+1):
			for j in range(1, sheet.max_column+1):
				if(i==1 and j>1):
					continue
				else:
					tickets.append(sheet.cell(row=i, column=j).value)

## Get all information of a table
	def getTable(self, urlStart,timeStart,timeEnd,urlEnd):	
		regex = re.compile(r'>(.*?)<')
		res = req.get(urlStart+timeStart+timeEnd+urlEnd, auth=(self.keys[0], self.keys[1]))
		res.raise_for_status()
		soup = bs4.BeautifulSoup(res.text, "lxml")
		table = soup.select('table.data')
		list = regex.findall(str(table[0]))
		while '' in list: list.remove('')
		return list

## TODO: Create Word document


## TODO: Create PDF document


## TODO: Confirm send to clients from email


## TODO: Send from email



start = [1, 12, 2017, 00, 00]
end = [31, 12, 2017, 00, 00]
bot = ReportBot()
bot.readHosts()
for i in range(0, len(bot.hosts)):
	print(bot.hosts[i])
	bot.takeData(bot.hosts[i], start, end)
print('Bot Stoped.')